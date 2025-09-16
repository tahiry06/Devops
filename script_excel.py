import pandas as pd
import os
import re
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import Tk, filedialog

def nettoyer_nom(valeur):
    nom = str(valeur)
    nom = re.sub(r'[\\/:*?"<>|]', '-', nom)
    nom = nom.replace(' ', '_').strip()[:50]
    return nom or "departement_inconnu"

# -- Ouvrir dialogues pour choisir fichiers
Tk().withdraw()

# Choix du fichier source .xlsb
fichier_source = filedialog.askopenfilename(
    title="Sélectionner le fichier .xlsb",
    filetypes=[("Excel Binary Workbook", "*.xlsb")]
)
if not fichier_source:
    print("❌ Aucun fichier sélectionné.")
    exit()

# Choix du dossier de destination
dossier_cibles = filedialog.askdirectory(
    title="Sélectionner le dossier de destination"
)
if not dossier_cibles:
    print("❌ Aucun dossier sélectionné.")
    exit()

# Lire les données depuis le fichier source
feuille_principale = "Employee_liste"
try:
    df_source = pd.read_excel(fichier_source, sheet_name=feuille_principale, engine='pyxlsb', skiprows=1)
except Exception as e:
    print(f"❌ Erreur lecture : {e}")
    exit()

# Vérification de la colonne Département (colonne G, index 6)
col_index_dept = 6
if col_index_dept >= len(df_source.columns):
    print("❌ Colonne G (Département) absente.")
    exit()

col_dept = df_source.columns[col_index_dept]
df_source[col_dept] = df_source[col_dept].astype(str).str.strip()
df_source = df_source[df_source[col_dept] != ""]

# Définir les colonnes utilisées pour détecter les doublons
col_matricule = "Employee"
col_mois = "Mois"  # Remplace par le vrai nom de ta colonne "mois" si différent

if col_matricule not in df_source.columns or col_mois not in df_source.columns:
    print("❌ Colonnes 'Matricule' et/ou 'Mois' manquantes.")
    exit()

# Parcourir chaque département
for dept in df_source[col_dept].unique():
    df_filtré = df_source[df_source[col_dept] == dept].copy()  # ✅ copie propre

    nom_fichier = f"Classeur_{nettoyer_nom(dept)}.xlsm"
    chemin_fichier = os.path.join(dossier_cibles, nom_fichier)

    # Charger ou créer le classeur
    if os.path.exists(chemin_fichier):
        wb = load_workbook(chemin_fichier, keep_vba=True)
        if 'Employee_liste' in wb.sheetnames:
            ws = wb['Employee_liste']
        else:
            ws = wb.create_sheet('Employee_liste')
            for c_idx, col in enumerate(df_filtré.columns, 1):
                ws.cell(row=1, column=c_idx, value=col)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employee_liste'
        for c_idx, col in enumerate(df_filtré.columns, 1):
            ws.cell(row=1, column=c_idx, value=col)

    # Lire les paires (Matricule, Mois) déjà présentes
    header = [cell.value for cell in ws[1]]
    try:
        matricule_col_idx = header.index(col_matricule) + 1
        mois_col_idx = header.index(col_mois) + 1
    except ValueError:
        print(f"❌ Colonnes {col_matricule} ou {col_mois} non trouvées dans {nom_fichier}.")
        continue

    paires_existantes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        pair = (str(row[matricule_col_idx - 1]), str(row[mois_col_idx - 1]))
        paires_existantes.add(pair)

    # Générer la clé pour les nouvelles données
    df_filtré["__key__"] = df_filtré.apply(
        lambda row: (str(row[col_matricule]), str(row[col_mois])), axis=1
    )
    df_ajouter = df_filtré[~df_filtré["__key__"].isin(paires_existantes)].drop(columns="__key__")

    # Ajouter les nouvelles lignes
    if not df_ajouter.empty:
        next_row = ws.max_row + 1
        for r in dataframe_to_rows(df_ajouter, index=False, header=False):
            for c_idx, val in enumerate(r, 1):
                ws.cell(row=next_row, column=c_idx, value=val)
            next_row += 1
        print(f"✅ {len(df_ajouter)} lignes ajoutées dans {nom_fichier}")
    else:
        print(f"⚠️ Aucune nouvelle ligne à ajouter pour {nom_fichier}")

    # Sauvegarde
    try:
        wb.save(chemin_fichier)
    except Exception as e:
        print(f"❌ Erreur sauvegarde fichier {nom_fichier} : {e}")

print("\n🎉 Fin du traitement sans doublons.")
